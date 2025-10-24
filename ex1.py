from pulp import *
import pandas as pd

def solve_with_pulp():
    #Создаем функцию и задаем задачу, название задачи, LpMaximize - максимизация ЦФ
    prob = LpProblem("Ballast_Production", LpMaximize)
    
    # Переменные решения, название, lowBound - неотрицательность (>=0), cat='Continuous' - непрерывные переменные
    x1 = LpVariable("x1", lowBound=0, cat='Continuous')# Объем 1-го типа балласта (песчаного)
    x2 = LpVariable("x2", lowBound=0, cat='Continuous')# Объем 2-го типа балласта (песчано-гравийного)
    x3 = LpVariable("x3", lowBound=0, cat='Continuous')# Объем 3-го типа балласта (щебеночного)
    
    # Целевая функция
    prob += 6*x1 + 10*x2 + 12*x3, "Total_Profit"
    
    # Ограничения
    prob += 13*x1 + 27*x2 + 24*x3 <= 230, "Excavators"
    prob += 8*x1 + 4*x2 + 6*x3 <= 50, "Bulldozers"
    prob += 50*x1 + 30*x2 + 50*x3 <= 610, "Labor"
    prob += x2 <= 8, "Demand_x2"
    prob += x3 <= 5, "Demand_x3"
    
    # Решение задачи
    prob.solve()
    
    # Вывод результатов в консоль
    print("=" * 50)
    print("РЕШЕНИЕ")
    print("=" * 50)
    print(f"Статус: {LpStatus[prob.status]}")
    print(f"Максимальная прибыль: {value(prob.objective):.2f} тыс. ден. ед.")
    print(f"\nОптимальные объемы:")
    for v in prob.variables():
        print(f"{v.name} = {v.varValue:.2f} тыс. м³")
    
    # Создание Excel документа
    create_excel_report(prob)
    
    return prob

def create_excel_report(prob):
    """Создание Excel ведомости в формате исходного документа"""
    
    # Получаем результаты
    results = {}
    for v in prob.variables():
        results[v.name] = v.varValue
    
    # Создаем данные для таблицы
    data = []
    
    # Заголовок таблицы
    data.append(['Ведомость объема работ'])
    data.append([])  # Пустая строка
    
    # Шапка таблицы
    data.append(['№ пп', 'Наименование', 'Объем работ', '', 'Стоимость работ', ''])
    data.append(['', '', 'Ед.изм.', 'Кол-во', 'Ед.изм.', 'Кол-во'])
    data.append(['1', '2', '3', '4', '5', '6'])
    
    # Данные по балласту
    data.append([
        '1',
        'Добыча и производство песчаного балласта',
        'м³',
        f"{results['x1']:.2f}",
        'тыс.ден.ед',
        f"{6 * results['x1']:.2f}"
    ])
    
    data.append([
        '2',
        'Добыча и производство песчано-гравийного балласта',
        'м³',
        f"{results['x2']:.2f}",
        'тыс.ден.ед',
        f"{10 * results['x2']:.2f}"
    ])
    
    data.append([
        '3',
        'Добыча и производство щебеночного балласта',
        'м³',
        f"{results['x3']:.2f}",
        'тыс.ден.ед',
        f"{12 * results['x3']:.2f}"
    ])
    
    # Итоги
    #total_volume = sum(results.values())
    total_cost = 6*results['x1'] + 10*results['x2'] + 12*results['x3']
    data.append(['Итого', '', '', '', '', f"{total_cost:.2f}"])
    
    # Пустые строки перед подписью
    data.append([])
    data.append([])
    
    # Подпись
    data.append(['', '', '',  'Составил:','', 'Романова О.А.'])
    
    # Создаем DataFrame
    df = pd.DataFrame(data)
    
    # Создаем Excel файл
    filename = f"VolumeStatement.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Ведомость', index=False, header=False)
        
        # Получаем workbook и worksheet для форматирования
        workbook = writer.book
        worksheet = writer.sheets['Ведомость']
        
        # Настраиваем ширину колонок
        column_widths = [8, 50, 10, 14, 14, 14]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[chr(64 + i)].width = width
        
        # Форматируем заголовок
        from openpyxl.styles import Font, Alignment, Border, Side
        
        # Стиль для заголовка
        title_font = Font(size=16, bold=True)
        worksheet['A1'].font = title_font
        worksheet.merge_cells('A1:F1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # Стиль для шапки таблицы
        header_font = Font(bold=True)
        for row in range(3, 6):  # Строки с заголовками
            for col in range(1, 7):
                cell = worksheet.cell(row=row, column=col)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Объединяем ячейки шапки
        worksheet.merge_cells('A3:A4')  # № пп
        worksheet.merge_cells('B3:B4')  # Наименование
        worksheet.merge_cells('C3:D3')  # Объем работ
        worksheet.merge_cells('E3:F3')  # Стоимость работ
        worksheet.merge_cells('A9:C9')  # Итого
        
        # Границы для таблицы
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        
        # Применяем границы ко всей таблице
        for row in range(3, 10):  # От заголовков до итогов
            for col in range(1, 7):
                worksheet.cell(row=row, column=col).border = thin_border
        
        # Выравнивание для числовых данных
        for row in range(6, 9):  # Строки с данными
            worksheet.cell(row=row, column=4).alignment = Alignment(horizontal='right')
            worksheet.cell(row=row, column=6).alignment = Alignment(horizontal='right')
        
        # Итоговая строка
        worksheet.cell(row=12, column=4).alignment = Alignment(horizontal='right')
        worksheet.cell(row=12, column=6).alignment = Alignment(horizontal='right')
        worksheet.cell(row=12, column=1).font = Font(bold=True)
        
        # Подпись
        worksheet.cell(row=14, column=5).font = Font(bold=True)
        worksheet.cell(row=15, column=5).font = Font(bold=True)
    
        print(f"\nExcel ведомость сохранена как: {filename}")

# Запуск решения
if __name__ == "__main__":
    solve_with_pulp()