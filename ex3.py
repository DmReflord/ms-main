from pulp import *
import pandas as pd

def create_assignment_matrix(prob):
    """Создает матрицу распределения 0 и 1"""
    
    # Создаем пустую матрицу 4x4
    assignment_matrix = [[0] * 4 for _ in range(4)]
    
    # Заполняем матрицу значениями из решения
    for i in range(4):
        for j in range(4):
            var_name = f"x{i+1}{j+1}"
            var_value = prob.variablesDict()[var_name].varValue
            assignment_matrix[i][j] = int(var_value)  # Преобразуем в 0 или 1
    
    return assignment_matrix

def print_assignment_matrix(assignment_matrix):
    """Выводит матрицу распределения в консоль"""
    
    print("\nМАТРИЦА РАСПРЕДЕЛЕНИЯ:")
    print("       Объект1  Объект2  Объект3  Объект4")
    print("      " + "-" * 35)
    
    for i in range(4):
        row_str = f"Бриг{i+1} |"
        for j in range(4):
            row_str += f"    {assignment_matrix[i][j]}     "
        print(row_str)

def solve_assignment_compact():
    prob = LpProblem("Brigade_Assignment", LpMinimize)
    
    # Матрица времени
    time_matrix = [
        [30, 40, 50, 60],  # Бригада 1
        [37, 47, 57, 58],  # Бригада 2
        [27, 44, 49, 57],  # Бригада 3
        [35, 37, 47, 63]   # Бригада 4
    ]
    
    # Создаем переменные
    x = [[LpVariable(f"x{i+1}{j+1}", cat='Binary') for j in range(4)] for i in range(4)]
    
    # Целевая функция
    prob += lpSum(time_matrix[i][j] * x[i][j] for i in range(4) for j in range(4))
    
    # Ограничения
    for i in range(4):  # Каждая бригада на одном объекте
        prob += lpSum(x[i][j] for j in range(4)) == 1
    
    for j in range(4):  # На каждый объект одна бригада
        prob += lpSum(x[i][j] for i in range(4)) == 1
    
    # Решение
    prob.solve()
    
    print("\n" + "=" * 60)
    print("РЕЗУЛЬТАТЫ РАСПРЕДЕЛЕНИЯ БРИГАД")
    print("=" * 60)
    print(f"Статус: {LpStatus[prob.status]}")
    print(f"Минимальное суммарное время: {value(prob.objective)} дней")
    
    # СОЗДАЕМ МАТРИЦУ РАСПРЕДЕЛЕНИЯ
    assignment_matrix = create_assignment_matrix(prob)
    
    # ВЫВОДИМ МАТРИЦУ
    print_assignment_matrix(assignment_matrix)
    
    # Собираем результаты назначений
    assignments = []
    print("\nНазначения:")
    for i in range(4):
        for j in range(4):
            if x[i][j].varValue == 1:
                print(f"Бригада {i+1} → Объект {j+1} (время: {time_matrix[i][j]} дней)")
                assignments.append({
                    'brigade': i+1,
                    'object': j+1,
                    'time': time_matrix[i][j]
                })
    
    # Создание Excel документа
    create_excel_report(assignments)
    
    return prob, assignments

def create_excel_report(assignments):
    """Создание Excel ведомости в точном формате"""
    
    # Создаем данные для таблицы
    data = []
    
    # Заголовок таблицы
    data.append(['Ведомость распределения бригад по объектам строительства'])
    data.append([])  # Пустая строка
    
    # Шапка таблицы - ПРАВИЛЬНАЯ СТРУКТУРА
    data.append(['', '', '', 'Срок', ''])  # Верхняя строка шапки
    data.append(['№ пп', 'Бригада', 'Объект', 'Ед.изм.', 'Кол-во'])  # Основная шапка
    data.append(['1', '2', '3', '4', '5'])  # Нумерация колонок
    
    # Данные по назначениям
    total_time = 0
    for idx, assignment in enumerate(assignments, 1):
        data.append([
            idx,
            assignment['brigade'],
            assignment['object'],
            'дни',
            assignment['time']
        ])
        total_time += assignment['time']
    
    # Итоги
    data.append(['Итого', '', '', '', total_time])
    
    # Пустые строки перед подписью
    data.append([])
    data.append([])
    
    # Подпись
    data.append(['', '', '', 'Составил:', 'Романова О.А.'])
    
    # Создаем DataFrame
    df = pd.DataFrame(data)
    
    # Создаем Excel файл
    filename = "Brigade_Assignment_Report.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Ведомость', index=False, header=False)
        
        # Получаем workbook и worksheet для форматирования
        workbook = writer.book
        worksheet = writer.sheets['Ведомость']
        
        # Настраиваем ширину колонок
        column_widths = [8, 11, 11, 14, 14]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[chr(64 + i)].width = width
        
        # Форматируем заголовок
        from openpyxl.styles import Font, Alignment, Border, Side
        
        # Стиль для заголовка
        title_font = Font(size=11, bold=True) 
        worksheet['A1'].font = title_font
        worksheet.merge_cells('A1:E1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # Стиль для шапки таблицы
        header_font = Font(bold=True)
        
        # Объединяем ячейки шапки ПРАВИЛЬНО
        worksheet.merge_cells('D3:E3')  # Объединяем "Срок" по горизонтали
        
        # Применяем стиль ко всем строкам шапки
        for row in [3, 4, 5]:
            for col in range(1, 6):
                cell = worksheet.cell(row=row, column=col)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Границы для таблицы
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        
        # Применяем границы ко всей таблице (от шапки до итогов)
        start_row = 3  # Начало таблицы (первая строка шапки)
        end_row = 5 + len(assignments) + 1  # До строки с итогами включительно
        
        for row in range(start_row, end_row + 1):
            for col in range(1, 6):
                if row <= worksheet.max_row and col <= 5:
                    worksheet.cell(row=row, column=col).border = thin_border
        
        # ВЫРАВНИВАНИЕ ПО ЦЕНТРУ для всех данных
        for row in range(6, 6 + len(assignments)):  # Строки с данными
            for col in range(1, 6):  # Все колонки
                worksheet.cell(row=row, column=col).alignment = Alignment(
                    horizontal='center', 
                    vertical='center'
                )
        
        # Итоговая строка
        total_row = 6 + len(assignments)
        worksheet.cell(row=total_row, column=1).font = Font(bold=True)
        worksheet.cell(row=total_row, column=5).font = Font(bold=True)
        worksheet.cell(row=total_row, column=5).alignment = Alignment(
            horizontal='center', 
            vertical='center'
        )
        
        # Объединяем ячейки для "Итого"
        worksheet.merge_cells(f'A{total_row}:D{total_row}')
        
        # Подпись
        signature_row = total_row + 3
        worksheet.cell(row=signature_row, column=4).font = Font(bold=True)
        worksheet.cell(row=signature_row, column=5).font = Font(bold=True)
        worksheet.cell(row=signature_row, column=4).alignment = Alignment(horizontal='right')
        worksheet.cell(row=signature_row, column=5).alignment = Alignment(horizontal='center')
        
        # Выравнивание для пустых ячеек в подписи
        for col in [1, 2, 3]:
            worksheet.cell(row=signature_row, column=col).alignment = Alignment(horizontal='center')
    
    print(f"\nExcel ведомость сохранена как: {filename}")
    
    return filename

# Запуск решения
if __name__ == "__main__":
    prob, assignments = solve_assignment_compact()