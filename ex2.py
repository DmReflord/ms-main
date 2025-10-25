from pulp import *
import pandas as pd

def solve_ex2():
    
    #Создаем функцию и задаем задачу, название задачи, LpMinimize - минимизация ЦФ
    prob = LpProblem("Ballast_Traffic", LpMinimize)
    
    # Переменные решения, название, lowBound - неотрицательность (>=0), cat='Continuous' - непрерывные переменные
    x11 = LpVariable("x11", lowBound=0, cat='Continuous')  # Объем балласта 1-го карьера на 1-ый участок 
    x12 = LpVariable("x12", lowBound=0, cat='Continuous')  # Объем балласта 1-го карьера на 2-ый участок 
    x21 = LpVariable("x21", lowBound=0, cat='Continuous')  # Объем балласта 2-го карьера на 1-ый участок
    x22 = LpVariable("x22", lowBound=0, cat='Continuous')  # Объем балласта 2-го карьера на 2-ый участок
    
    # Целевая функция
    prob += 10*x11 + 9*x12 + 4*x21 + 5*x22, "Total_Profit"
    
    # Ограничения
    prob += x11 + x12 <= 35, "Power first provider"
    prob += x21 + x22 <= 25, "Power second provider"
    prob += x11 + x21 >= 27, "Demand first field"
    prob += x12 + x22 >= 15, "Demand second field"
    
    # Решение задачи
    prob.solve()
    
    # Вывод результатов
    print("=" * 50)
    print("РЕШЕНИЕ")
    print("=" * 50)
    print(f"Статус: {LpStatus[prob.status]}")
    print(f"Минимальные затраты: {value(prob.objective):.2f} тыс. ден. ед.")
    print(f"\nОптимальные объемы:")
    for v in prob.variables():
        print(f"{v.name} = {v.varValue:.2f} тыс. м³")
     # Создание Excel документа
    create_excel_report(prob)
    
    return prob
    
def create_excel_report(prob):
    """Создание Excel ведомости в формате исходного документа"""
    
    # Получаем результаты
    results = {}
    for v in prob.variables():
        results[v.name] = v.varValue
    
    # Создаем данные для таблицы
    data = []
    
    # Заголовок таблицы
    data.append(['Ведомость объема работ'])
    data.append([])  # Пустая строка
    
    # Шапка таблицы
    data.append(['№ пп', 'Наименование', 'Поставщик', 'Потребитель','Объем работ', '', 'Затраты', ''])
    data.append(['', '','', '', 'Ед.изм.', 'Кол-во', 'Ед.изм.', 'Кол-во'])
    data.append(['1', '2', '3', '4', '5', '6','7','8'])
    
    # Данные по балласту
    data.append([
        '1',
        'Балласт',
        '1',
        '1',
        'м³',
        f"{results['x11']:.2f}",
        'тыс.ден.ед',
        f"{10 * results['x11']:.2f}"
    ])
    
    data.append([
        '2',
        'Балласт',
        '1',
        '2',
        'м³',
        f"{results['x12']:.2f}",
        'тыс.ден.ед',
        f"{9* results['x12']:.2f}"
    ])
    
    data.append([
        '3',
        'Балласт',
        '2',
        '1',
        'м³',
        f"{results['x21']:.2f}",
        'тыс.ден.ед',
        f"{4 * results['x21']:.2f}"
    ])

    data.append([
        '4',
        'Балласт',
        '2',
        '2',
        'м³',
        f"{results['x22']:.2f}",
        'тыс.ден.ед',
        f"{5 * results['x22']:.2f}"
    ])
    
    # Итоги
    #total_volume = sum(results.values())
    total_cost = 10*results['x11'] + 9*results['x12'] + 4*results['x21'] + 5*results['x22']
    data.append(['Итого', '', '', '', '','', '', f"{total_cost:.2f}"])
    
    # Пустые строки перед подписью
    data.append([])
    data.append([])
    
    # Подпись
    data.append(['', '', '',  'Составил:','', 'Романова О.А.'])
    
    # Создаем DataFrame
    df = pd.DataFrame(data)
    
    # Создаем Excel файл
    filename = f"VolumeStatement_2.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Ведомость', index=False, header=False)
        
        # Получаем workbook и worksheet для форматирования
        workbook = writer.book
        worksheet = writer.sheets['Ведомость']
        
        # Настраиваем ширину колонок
        column_widths = [8, 25, 15, 15, 14, 14, 14, 14]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[chr(64 + i)].width = width
        
        # Форматируем заголовок
        from openpyxl.styles import Font, Alignment, Border, Side
        
        # Стиль для заголовка
        title_font = Font(size=16, bold=True)
        worksheet['A1'].font = title_font
        worksheet.merge_cells('A1:F1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # Стиль для шапки таблицы
        header_font = Font(bold=True)
        for row in range(3, 10):  # Строки с заголовками
            for col in range(1, 9):
                cell = worksheet.cell(row=row, column=col)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Объединяем ячейки шапки
        worksheet.merge_cells('A3:A4')  # № пп
        worksheet.merge_cells('B3:B4')  # Наименование
        worksheet.merge_cells('C3:C4')  # Поставщик
        worksheet.merge_cells('D3:D4')  # Потребитель
        worksheet.merge_cells('E3:F3')  # Объем работ
        worksheet.merge_cells('G3:H3')  # Затраты
        worksheet.merge_cells('A10:G10')  # Итого
        
        # Границы для таблицы
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        
        # Применяем границы ко всей таблице
        for row in range(3, 11):  # От заголовков до итогов
            for col in range(1, 9):
                worksheet.cell(row=row, column=col).border = thin_border
        
        # Выравнивание для числовых данных
        for row in range(6, 11):  # Строки с данными
            worksheet.cell(row=row, column=6).alignment = Alignment(horizontal='right')
            worksheet.cell(row=row, column=8).alignment = Alignment(horizontal='right')
        
        # Итоговая строка
        worksheet.cell(row=10, column=6).alignment = Alignment(horizontal='right')
        worksheet.cell(row=10, column=8).alignment = Alignment(horizontal='right')
        worksheet.cell(row=10, column=1).font = Font(bold=True)
        
        # Подпись
        worksheet.cell(row=13, column=4).font = Font(bold=True)
        worksheet.cell(row=13, column=6).font = Font(bold=True)
    
        print(f"\nExcel ведомость сохранена как: {filename}")

# Запуск решения
if __name__ == "__main__":
    solve_ex2()