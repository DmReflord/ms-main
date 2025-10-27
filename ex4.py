import numpy as np
from pulp import *
import pandas as pd

def ex_4():

    # Данные из контрольного примера
    count = [0, 17, 34, 51, 68]  # Количество рабочих

    CMR = np.array([
        [0,  0,  0,  0],
        [8,  9,  7, 6],
        [14, 16, 16, 10],
        [24, 25, 22, 18],
        [32, 33, 30, 24]
    ])

    # Бинарные переменные
    v = np.array([
        [LpVariable("v11", cat='Binary'), LpVariable("v12", cat='Binary'), 
        LpVariable("v13", cat='Binary'), LpVariable("v14", cat='Binary')],
        [LpVariable("v21", cat='Binary'), LpVariable("v22", cat='Binary'), 
        LpVariable("v23", cat='Binary'), LpVariable("v24", cat='Binary')],
        [LpVariable("v31", cat='Binary'), LpVariable("v32", cat='Binary'), 
        LpVariable("v33", cat='Binary'), LpVariable("v34", cat='Binary')],
        [LpVariable("v41", cat='Binary'), LpVariable("v42", cat='Binary'), 
        LpVariable("v43", cat='Binary'), LpVariable("v44", cat='Binary')],
        [LpVariable("v51", cat='Binary'), LpVariable("v52", cat='Binary'), 
        LpVariable("v53", cat='Binary'), LpVariable("v54", cat='Binary')]
    ])

    C = 68  # Всего рабочих

    problem = LpProblem('Maximize_CMR', LpMaximize)

    # Целевая функция
    profit = lpSum(CMR[i][j] * v[i][j] for i in range(5) for j in range(4))
    problem += profit

    # Ограничение на общее количество рабочих
    problem += (lpSum(count[i] * v[i][j] for i in range(5) for j in range(4)) == C)

    # Каждому объекту назначается ровно одна группа рабочих
    for j in range(4):
        problem += lpSum(v[i][j] for i in range(5)) == 1

    # Решение
    status = problem.solve()

    print("Матрица распределения:")
    print("=" * 50)
    print("Объекты →", end=" ")
    for j in range(4):
        print(f"  {j+1}  ", end=" ")
    print("\n" + "=" * 50)

    for i in range(5):
        print(f"Группа {i} ({count[i]} раб.) |", end=" ")
        for j in range(4):
            val = v[i][j].varValue 
            print(f" {val:5.1f} ", end=" ")
        print()

    print("=" * 50)
    print("Status:", LpStatus[status])
    print("Result:")
    for i in range(5):
        print(f"Группа {i}:", end=" ")
        for j in range(4):
            print(f"{v[i][j].varValue:>4}", end=" ")
        print()

    # Используем pulp.value() явно
    print("Максимальный объем СМР:", pulp.value(problem.objective))

    # Проверка распределения рабочих
    total_workers = 0
    total_cmr = 0
    assignments = []
    for j in range(4):
        for i in range(5):
            if v[i][j].varValue == 1:
                total_workers += count[i]
                total_cmr += CMR[i][j]
                print(f"Объект {j+1}: {count[i]} рабочих, СМР = {CMR[i][j]} тыс.руб")
                assignments.append({
                    'object': j+1,
                    'count': count[i],
                    'cmr': CMR[i][j]
                })

    print(f"Всего распределено рабочих: {total_workers}")
    print(f"Суммарный объем СМР: {total_cmr} тыс.руб")
    # Создание Excel документа
    create_excel_report(assignments)
    
    return problem, assignments

def create_excel_report(assignments):
    """Создание Excel ведомости в точном формате"""
    
    # Создаем данные для таблицы
    data = []
    
    # Заголовок таблицы
    data.append(['Ведомость распределения рабочих по объектам строительства'])
    data.append([])  # Пустая строка
    
    # Шапка таблицы 
    data.append(['№ пп', 'Объект', 'Количество рабочих','', 'Объем СМР',''])  # Основная шапка
    data.append(['', '', 'Ед.изм.','Кол-во', 'Ед.изм.','Кол-во'])
    data.append(['1', '2', '3', '4', '5','6'])  # Нумерация колонок
    
    # Данные по назначениям
    total_cmr = 0
    for idx, assignment in enumerate(assignments, 1):
        data.append([
            idx,
            assignment['object'],
            'чел',
            assignment['count'],
            'тыс.руб',
            assignment['cmr']
        ])
        total_cmr += assignment['cmr']
    
    # Итоги
    data.append(['Итого', '', '', '', '',total_cmr])
    
    # Пустые строки перед подписью
    data.append([])
    data.append([])
    
    # Подпись
    data.append(['', '', '', 'Составил:', 'Романова О.А.'])
    
    # Создаем DataFrame
    df = pd.DataFrame(data)
    
    # Создаем Excel файл
    filename = "Ex4.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Ведомость', index=False, header=False)
        
        # Получаем workbook и worksheet для форматирования
        workbook = writer.book
        worksheet = writer.sheets['Ведомость']
        
        # Настраиваем ширину колонок
        column_widths = [8, 12, 12, 14, 14]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[chr(64 + i)].width = width
        
        # Форматируем заголовок
        from openpyxl.styles import Font, Alignment, Border, Side
        
        # Стиль для заголовка
        title_font = Font(size=11, bold=True) 
        worksheet['A1'].font = title_font
        worksheet.merge_cells('A1:F1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # Стиль для шапки таблицы
        header_font = Font(bold=True)
        
        # Объединяем ячейки шапки ПРАВИЛЬНО
        worksheet.merge_cells('C3:D3')
        worksheet.merge_cells('A3:A4')
        worksheet.merge_cells('B3:B4')
        
        # Применяем стиль ко всем строкам шапки
        for row in [3, 4]:
            for col in range(1, 7):
                cell = worksheet.cell(row=row, column=col)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Границы для таблицы
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        
        # Применяем границы ко всей таблице (от шапки до итогов)
        start_row = 3  # Начало таблицы (первая строка шапки)
        end_row = 5 + len(assignments) + 1  # До строки с итогами включительно
        
        for row in range(start_row, end_row + 1):
            for col in range(1, 7):
                if row <= worksheet.max_row and col <= 6:
                    worksheet.cell(row=row, column=col).border = thin_border
        
        # ВЫРАВНИВАНИЕ ПО ЦЕНТРУ для всех данных
        for row in range(5, 5 + len(assignments) + 1):  # Строки с данными
            for col in range(1, 7):  # Все колонки
                worksheet.cell(row=row, column=col).alignment = Alignment(
                    horizontal='center', 
                    vertical='center'
                )
        
        # Итоговая строка
        total_row = 6 + len(assignments)
        worksheet.cell(row=total_row, column=1).font = Font(bold=True)
        worksheet.cell(row=total_row, column=5).font = Font(bold=True)
        worksheet.cell(row=total_row, column=5).alignment = Alignment(
            horizontal='center', 
            vertical='center'
        )
        
        # Объединяем ячейки для "Итого"
        worksheet.merge_cells(f'A{total_row}:D{total_row}')
        
        # Подпись
        signature_row = total_row + 3
        worksheet.cell(row=signature_row, column=4).font = Font(bold=True)
        worksheet.cell(row=signature_row, column=5).font = Font(bold=True)
        worksheet.cell(row=signature_row, column=4).alignment = Alignment(horizontal='right')
        worksheet.cell(row=signature_row, column=5).alignment = Alignment(horizontal='center')
        
        # Выравнивание для пустых ячеек в подписи
        for col in [1, 2, 3]:
            worksheet.cell(row=signature_row, column=col).alignment = Alignment(horizontal='center')
    
    print(f"\nExcel ведомость сохранена как: {filename}")
    
    return filename

# Запуск решения
if __name__ == "__main__":
    problem, assignments = ex_4()